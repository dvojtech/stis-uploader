# stis_uploader.py
import argparse, os, re, sys, time, shutil
import unicodedata
import traceback
import ctypes
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
from playwright.sync_api import TimeoutError as PwTimeout

# --- rychlé timeouty (ms) pro výběr hráčů ---
FAST_CLICK_MS   = 400
FAST_FOCUS_MS   = 300
FAST_MENU_MS    = 700   # čekání na zobrazení autocomplete
FAST_PAUSE_MS   = 80
MAX_PER_NAME_MS = 1500  # tvrdý strop ~1.5 s na 1 jméno
AFTER_SELECT_SLEEP_MS = 60

DIAG_DIR = Path(os.getcwd()) / "stis_diag"
DIAG_DIR.mkdir(exist_ok=True)

# kde EXE skutečně leží
EXE_DIR = Path(sys.argv[0]).resolve().parent
# kam určitě umíme zapsat
TEMP_DIR = Path(os.environ.get("TEMP", str(EXE_DIR)))
# cesty pro "boot" log (zapisujeme na obě místa)
BOOT_FILES = [
    TEMP_DIR / "stis_boot.log",
    EXE_DIR / "stis_boot.log",
]
# --- konfigurace mapování řádků v listu "zdroj" ---
ZDROJ_SHEET             = "zdroj"
ZDROJ_FIRST_SINGLE_ROW  = 7     # první řádek singlů (D/E = jména, I—M = sety)
SINGLES_COUNT           = 16    # kolik singlů se vyplňuje (2..17)

from pathlib import Path
import os, sys, shutil
_A1COL = {c:i for i,c in enumerate("ABCDEFGHIJKLMNOPQRSTUVWXYZ", start=1)}
def a1_to_rc(a1: str):
    """
    'D2' -> (row=2, col=4)
    'I3' -> (3, 9)
    """
    a1 = a1.strip().upper()
    m = re.fullmatch(r"([A-Z]+)(\d+)", a1)
    if not m:
        raise ValueError(f"Bad A1 address: {a1}")
    col_s, row_s = m.groups()
    col = 0
    for ch in col_s:
        col = col * 26 + (ord(ch) - 64)
    return int(row_s), int(col)

def cell_value(sh, a1: str):
    r, c = a1_to_rc(a1)
    v = sh.cell(r, c).value
    return ("" if v is None else str(v)).strip()

def row_sets(sh, row: int, cols=("I","J","K","L","M")):
    vals = []
    for col in cols:
        v = cell_value(sh, f"{col}{row}")
        v = v.strip()
        if v == "":
            vals.append(None)  # prázdný set
        else:
            vals.append(v)
    # ořízni trailing None (I…M často nemají všech 5 setů)
    while vals and vals[-1] is None:
        vals.pop()
    return vals
def fill_leaders_on_start(page, home_name_text: str, away_name_text: str, log, only_from_club=True):
    """
    Vyplní 'Vedoucí družstev' s PODMÍNKOU PŘESNÉ SHODY CELÉHO TEXTU položky.
    - Viditelné inputy:  id_domaci_vedoucitext / id_hoste_vedoucitext
    - Skrytá ID:         id_domaci_vedouciid   / id_hoste_vedouciid
    - Checkboxy 'Jen z oddílu': name='chbklub' a 'chbklub2'
    - ŽÁDNÉ fallbacky na první položku, žádná fuzzy shoda.
    - Když shoda není, vrátí False a zaloguje seznam položek (pro korekci textu).
    """
    MENU_MS  = 1500
    TYPE_DLY = 20
    SLEEP_MS = 80

    def _ensure_only_from_club():
        if not only_from_club:
            return
        for cb_name in ("chbklub", "chbklub2"):
            try:
                cb = page.locator(f"input[name='{cb_name}']").first
                if cb.count() and not cb.is_checked():
                    cb.check(timeout=600)
                    log(f"  [leaders] '{cb_name}' zaškrtnuto")
            except Exception:
                pass
def fill_leaders_on_start(page, home_name_text: str, away_name_text: str, log, only_from_club=True):
    """
    Vyplní 'Vedoucí družstev' klikem na položku v autocomplete menu.
    Priorita shody:
      1) přesná shoda CELÉHO textu (když předáš display text přesně jako v menu),
      2) když ne, vybere první položku, jejíž text ZAČÍNÁ na zadané jméno (prefix match),
         např. 'Dvořák Jiří' → 'Dvořák Jiří 1970 (TJ ...)'.
    Po výběru provede blur (Tab) a ověří hidden ID (…vedouciid).
    """
    MENU_MS  = 1500
    TYPE_DLY = 15
    SLEEP_MS = 80

    def _ensure_only_from_club():
        if not only_from_club:
            return
        for cb_name in ("chbklub", "chbklub2"):
            try:
                cb = page.locator(f"input[name='{cb_name}']").first
                if cb.count() and not cb.is_checked():
                    cb.check(timeout=600)
                    log(f"  [leaders] '{cb_name}' zaškrtnuto")
            except Exception:
                pass

    def _pick_click(input_sel: str, hidden_sel: str, display_hint: str) -> bool:
        hint = (display_hint or "").strip()
        if not hint:
            log(f"  [leaders] požadovaný text je prázdný pro {input_sel}")
            return False

        inp = page.locator(input_sel).first
        hid = page.locator(hidden_sel).first
        if not inp.count():
            log(f"  [leaders] input {input_sel} nenalezen")
            return False

        # vynuluj hidden (ať víme, že se po výběru nastaví)
        try:
            if hid.count():
                hid.evaluate("el => { el.value=''; }")
        except Exception:
            pass

        # napiš dotaz a otevři menu
        try: inp.fill("")
        except Exception: pass
        inp.focus()
        page.keyboard.type(hint, delay=TYPE_DLY)
        try:
            inp.evaluate("""
                el => {
                  el.dispatchEvent(new Event('input',{bubbles:true}));
                  el.dispatchEvent(new KeyboardEvent('keyup',{key:' ',bubbles:true}));
                }
            """)
        except Exception:
            pass

        menu_sel = "ul.ui-autocomplete:visible, .ui-autocomplete.ui-menu:visible"
        try:
            page.wait_for_selector(menu_sel, timeout=MENU_MS)
        except Exception:
            try:
                inp.evaluate("el => { if (window.jQuery && jQuery.fn.autocomplete) jQuery(el).autocomplete('search', el.value||''); }")
                page.wait_for_selector(menu_sel, timeout=MENU_MS)
            except Exception:
                log(f"  [leaders] menu se neukázalo pro {hint!r}")
                return False

        menu = page.locator(menu_sel).first.locator("li")
        cnt = menu.count()
        if not cnt:
            log(f"  [leaders] prázdné menu pro {hint!r}")
            return False

        # načti všechny texty položek najednou
        items = [ (i, (menu.nth(i).inner_text() or "").strip()) for i in range(cnt) ]
        log("  [leaders] menu:", "; ".join([f"{i}:{t}" for i,t in items[:10]]))

        # 1) přesná shoda celého textu
        pick = next((i for i,t in items if t == hint), -1)

        # 2) prefix shoda (case/diakritika ignorována): „začíná na zadané jméno“
        if pick < 0:
            hint_norm = _norm_name(hint)
            for i,t in items:
                if _norm_name(t).startswith(hint_norm):
                    pick = i; break

        if pick < 0:
            log(f"  [leaders] nenašla se shoda pro {hint!r} → nevybráno")
            return False

        # klik na vybranou li + blur
        try:
            menu.nth(pick).click(timeout=800)
        except Exception:
            page.keyboard.press("Enter")
        page.keyboard.press("Tab")
        page.wait_for_timeout(SLEEP_MS)

        # kontrola hidden ID i viditelného textu
        try:
            hid_val = (hid.get_attribute("value") or "").strip() if hid.count() else ""
        except Exception:
            hid_val = ""
        try:
            vis_val = inp.input_value(timeout=300)
        except Exception:
            vis_val = (inp.evaluate("el => el.value") or "").strip()

        ok = bool(hid_val) and bool(vis_val)
        log(f"  [leaders] '{hint}' → {'OK' if ok else 'NEULOŽENO'} (hidden={hid_val or '∅'}, visible={vis_val or '∅'})")
        return ok

    _ensure_only_from_club()
    ok_home = _pick_click("input[name='id_domaci_vedoucitext']",
                          "input[name='id_domaci_vedouciid']",
                          home_name_text)
    ok_away = _pick_click("input[name='id_hoste_vedoucitext']",
                          "input[name='id_hoste_vedouciid']",
                          away_name_text)
    return ok_home and ok_away


def fill_playroom(page, wanted_text: str, log):
    """
    Hrací místnost – robustně a „klikově“ přes JS:
      - najde <select name='zapis_id_herna'> (případně nejlepší kandidát),
      - zjistí seznam option (text + value),
      - vybere buď přesnou shodu podle textu, nebo první reálnou položku (index>0),
      - NÁSILNĚ nastaví selectedIndex a vystřelí input/change/click,
      - propíše vybraný text do input[name='zapis_herna'] a vystřelí jeho input/change.
    """
    CLICK_MS = 600
    SLEEP_MS = 80
    wanted_text = (wanted_text or "").strip()

    def _is_placeholder_text(t: str) -> bool:
        s = (t or "").strip().lower()
        return (s.startswith("- zvolte") or s.startswith("- vyberte") or s == "-" or "hrací místnost" in s)

    # 1) pokus o „oficiální“ select
    sel = page.locator("select[name='zapis_id_herna']").first
    if not sel.count():
        # fallbacky poblíž textu 'Hrací místnost'
        sel = page.locator("xpath=//*[contains(normalize-space(.),'Hrací místnost')]/following::select[1]").first
    if not sel.count():
        # nouzově první select, který nevypadá jako hodiny/minuty
        for i in range(min(page.locator("select").count(), 10)):
            cand = page.locator("select").nth(i)
            nm = (cand.get_attribute("name") or "").lower()
            if "hodin" in nm or "minut" in nm:
                continue
            sel = cand
            break

    if not sel.count():
        log("  [playroom] <select> pro 'Hrací místnost' nenalezen")
        return False

    # 2) načti seznam možností
    options = sel.evaluate("el => Array.from(el.options).map((o,i) => ({i, v:o.value, t:(o.textContent||'').trim()}))") or []
    if not options:
        log("  [playroom] select nemá položky")
        return False

    # 3) vyber index – nejprve přesná shoda podle TEXTU, potom první reálná (index>0, ne placeholder)
    pick_idx = -1
    if wanted_text:
        want_norm = _norm_name(_strip_menu_text(wanted_text))
        for o in options:
            if _norm_name(_strip_menu_text(o.get("t",""))) == want_norm:
                pick_idx = int(o["i"]); break

    if pick_idx < 0:
        for o in options:
            i = int(o["i"]); t = o.get("t","")
            if i == 0:            # přeskoč 0
                continue
            if _is_placeholder_text(t):
                continue
            pick_idx = i; break

    # poslední záchrana – když aspoň 2 položky, ber index 1
    if pick_idx < 0 and len(options) > 1:
        pick_idx = 1

    if pick_idx < 0:
        log("  [playroom] žádná vhodná položka k výběru (po všech pokusech)")
        log("  [playroom] options:", "; ".join([f"{o['i']}:{o['t']}" for o in options[:6]]))
        return False

    # 4) „klikově“ přes JS: nastav selectedIndex, vystřel input/change/click
    try:
        try: sel.scroll_into_view_if_needed(timeout=CLICK_MS)
        except Exception: pass
        try: sel.click(timeout=CLICK_MS, force=True)   # nevadí, když nic „nevyskočí“
        except Exception: pass

        sel.evaluate(
            """
            (el, idx) => {
              // nastav vybraný index
              el.selectedIndex = idx;

              // vystřel události, na které UI může poslouchat
              el.dispatchEvent(new Event('input',  {bubbles:true}));
              el.dispatchEvent(new Event('change', {bubbles:true}));
              el.dispatchEvent(new MouseEvent('click', {bubbles:true}));

              // propsat text do textového pole 'zapis_herna'
              const opt = el.options[el.selectedIndex];
              const txt = document.querySelector("input[name='zapis_herna']");
              if (opt && txt) {
                const label = (opt.textContent || '').trim();
                txt.value = label;
                txt.dispatchEvent(new Event('input',  {bubbles:true}));
                txt.dispatchEvent(new Event('change', {bubbles:true}));
              }
            }
            """,
            pick_idx
        )

        page.wait_for_timeout(SLEEP_MS)

        # 5) ověř – přečti skutečně vybraný text a zkontroluj, že to není placeholder
        chosen = sel.evaluate("el => (el.selectedOptions?.[0]?.textContent || el.options[el.selectedIndex]?.textContent || '').trim()") or ""
        log(f"  [playroom] vybráno: index={pick_idx}, text='{chosen}'")
        if _is_placeholder_text(chosen):
            log("  [playroom] FAIL: vybrán placeholder")
            return False
        return True

    except Exception as e:
        log(f"  [playroom] selhání: {e!r}")
        return False



def prepare_playwright_browsers(logger):
    """
    Najde přibalené ms-playwright (v _MEIPASS) a jednorázově ho zkopíruje
    vedle EXE (trvalá cesta). Potom nastaví proměnné tak, aby Playwright
    používal právě tuto kopii a automaticky nic nestahoval.
    """
    # adresář, kde běží EXE (nebo .py)
    exe_dir = Path(getattr(sys, "frozen", False) and sys.executable or __file__).resolve().parent
    # základna s přibalenými daty při PyInstaller onefile
    bundled_base = Path(getattr(sys, "_MEIPASS", exe_dir))
    bundled = bundled_base / "ms-playwright"
    # trvalá kopie vedle EXE
    persistent = exe_dir / "ms-playwright"

    # jednorázové kopírování přibalených dat vedle EXE
    if bundled.exists() and not persistent.exists():
        logger(f"Copying bundled ms-playwright → {persistent}")
        shutil.copytree(bundled, persistent)

    # preferuj trvalou kopii; když není, použij přímo přibalenou
    root = persistent if persistent.exists() else bundled

    # nastav prostředí a zakaž automatické stahování při běhu
    os.environ["PLAYWRIGHT_BROWSERS_PATH"] = str(root.resolve())
    os.environ["PW_DISABLE_DOWNLOADS"] = "1"
    os.environ.pop("PLAYWRIGHT_DOWNLOAD_HOST", None)

    # log info – najdi chrome(.exe) kdekoli pod root
    chromes = list(root.rglob("chrome.exe")) + list(root.rglob("chrome"))
    if chromes:
        logger(f"Using ms-playwright at: {root}")
    else:
        logger(f"WARNING: {root} neobsahuje žádný chrome(.exe) – "
               f"zvaž build s přibalenými prohlížeči nebo jednorázovou instalaci.")
    return root


def _norm_name(s: str) -> str:
    # normalizace jména: zmenší, odstraní diakritiku, srazí vícenásobné mezery
    s = " ".join((s or "").strip().split()).lower()
    s = "".join(ch for ch in unicodedata.normalize("NFD", s)
                if unicodedata.category(ch) != "Mn")
    return s

def _name_variants(full: str):
    # vrátí varianty "Jméno Příjmení" i "Příjmení Jméno"
    parts = [p for p in (full or "").strip().split() if p]
    if len(parts) >= 2:
        fn = " ".join(parts[:-1]); ln = parts[-1]
        return [f"{fn} {ln}", f"{ln} {fn}"]
    return [full]

def _strip_menu_text(t: str) -> str:
    # položky menu bývají "Příjmení Jméno (YYYY, Klub...)"
    # bereme text před závorkou/čárkou/pomlčkou
    base = re.split(r"[\(\[\-\,]", (t or "").strip(), maxsplit=1)[0].strip()
    return " ".join(base.split())

def _use_bundled_ms_playwright(log):
    """
    Když je EXE postavené s --add-data "ms-playwright;ms-playwright",
    nastavíme PLAYWRIGHT_BROWSERS_PATH na rozbalenou složku.
    Vrací True, pokud jsme cestu nastavili.
    """
    try:
        base = getattr(sys, "_MEIPASS", None) or os.path.dirname(sys.executable)
        cand = os.path.join(base, "ms-playwright")
        if os.path.isdir(cand):
            os.environ["PLAYWRIGHT_BROWSERS_PATH"] = cand
            log("Using bundled ms-playwright at:", cand)
            return True
    except Exception as e:
        log("Bundled ms-playwright detection failed:", repr(e))
    return False

def wait_online_ready(page, log):
    # STIS online editor: nahoře jsou tlačítka "Uložit změny" / "Dokončit zápis"
    sel_ready = "button:has-text('Uložit změny'), input[type='button'][value*='Uložit změny'], input[type='submit'][value*='Uložit změny']"
    page.wait_for_selector(sel_ready, timeout=30000)
    # pro jistotu ověř, že jsme na online.php
    if "online.php" not in page.url:
        log("Pozn.: nejsem na online.php, aktuální URL:", page.url)
    # informační log – kolik je editovatelných polí
    cnt_inputs = page.locator("#zapis input[type='text'], #zapis input[type='number']").count()
    log("Editor ready – editačních polí v #zapis:", cnt_inputs)



def _normalize_player_selector(selector: str) -> str:
    s = selector
    s = re.sub(r"(\.player(?:\.[a-z0-9_-]+)*)\s*\.player-name\b", r"\1", s, flags=re.I)
    s = re.sub(r"(\.player(?:\.[a-z0-9_-]+)*)\s*\.player\b", r"\1", s, flags=re.I)
    s = re.sub(r"\.player-name\b", ".player", s, flags=re.I)
    return s


def _any_visible_input(page):
    return page.locator("input[type='text']:visible, input.ui-autocomplete-input:visible, input.ac_input:visible").first

def _any_input_in_zapis(page):
    return page.locator("#zapis input[type='text']:visible, #zapis input.ui-autocomplete-input:visible, #zapis input.ac_input:visible").first

def _diag_dump_cell(page, target, tag, log):
    try:
        ts = int(time.time()*1000)
        cell_png = DIAG_DIR / f"{tag}_cell_{ts}.png"
        page_png = DIAG_DIR / f"{tag}_zapis_{ts}.png"
        html_snip = DIAG_DIR / f"{tag}_snippet_{ts}.html"

        page.locator("#zapis").first.screenshot(path=str(page_png))
        target.screenshot(path=str(cell_png))

        ae = page.evaluate("() => document.activeElement ? document.activeElement.outerHTML : null")
        outer = target.evaluate("el => el.outerHTML")  # ← klíčová změna

        with open(html_snip, "w", encoding="utf-8") as f:
            f.write("<h3>activeElement</h3><pre>")
            f.write(ae or "NULL")
            f.write("</pre><h3>cell outerHTML</h3>")
            f.write(outer)

        log(f"  [diag] uložené: {cell_png.name}, {page_png.name}, {html_snip.name}")
    except Exception as e:
        log(f"  [diag] dump selhal: {e!r}")

        
def _fill_player_by_click(page, selector, name, log):
    """
    BLESK výběr hráče:
    1) Primárně <select class="player"> v buňce – options si vezmu najednou přes evaluate.
    2) Jen když select není, zkusím rychlý autocomplete.
    Vše POUZE uvnitř hráčské buňky (žádné sety).
    """
    name = (name or "").strip()
    if not name:
        return

    # krátké defaulty (když nejsou definované globálně)
    MENU_MS  = globals().get("FAST_MENU_MS", 700)
    CLICK_MS = globals().get("FAST_CLICK_MS", 400)
    SLEEP_MS = globals().get("AFTER_SELECT_SLEEP_MS", 50)

    # --- odvoď selektor BUŇKY z dodaného selectoru ---
    cell_sel = None
    try:
        if " .cell-player:first-child" in selector or " .cell-player:last-child" in selector:
            cell_sel = selector.split(" .player", 1)[0]
        elif "#d" in selector:
            if ".player.domaci " in selector:
                cell_sel = selector.split(" .player", 1)[0].replace(".player.domaci", ".cell-player:first-child")
            if ".player.host " in selector:
                cell_sel = selector.split(" .player", 1)[0].replace(".player.host",   ".cell-player:last-child")
    except Exception:
        cell_sel = None

    cell = page.locator(cell_sel).first if cell_sel else page.locator(selector).first
    if not cell.count():
        log(f"  ✗ {name} → Nenalezen element: {cell_sel or selector}")
        return

    # krátký log „před“
    try:
        before_txt = (cell.inner_text() or "").strip()
        log(f"  → {name!r} @ {cell_sel or selector}  [before='{before_txt}']")
    except Exception:
        pass

    # ---------- FAST SELECT PATH ----------
    sel = cell.locator("select.player").first
    if not sel.count():
        # jeden nenáročný klik, kdyby select vznikal až po kliku
        try:
            try: cell.scroll_into_view_if_needed(timeout=CLICK_MS)
            except Exception: pass
            cell.click(timeout=CLICK_MS, force=True)
            page.wait_for_timeout(40)
        except Exception:
            pass
        sel = cell.locator("select.player").first

    if sel.count():
        try:
            # vyčti všechny možnosti jedním JS voláním (žádné pomalé per-option dotazy)
            opts = sel.evaluate("el => Array.from(el.options).map(o => ({v:o.value, t:(o.textContent||'').trim()}))")
            # připrav normalizované varianty hledaného jména
            want_norms = {_norm_name(v) for v in _name_variants(name)}

            def pick_value_from_options(options):
                # 1) přesná shoda (normalizovaně, bez „(rok, klub…)“ šumu)
                for o in options:
                    if _norm_name(_strip_menu_text(o.get('t',''))) in want_norms:
                        return o.get('v')
                # 2) fallback: samotné příjmení
                parts = name.split()
                if len(parts) >= 2:
                    ln = _norm_name(parts[-1])
                    for o in options:
                        norm = _norm_name(_strip_menu_text(o.get('t','')))
                        if norm.endswith(" " + ln) or norm == ln:
                            return o.get('v')
                return None

            pick_val = pick_value_from_options(opts)
            if not pick_val:
                log("  žádná shoda v <select> – přeskočeno")
                return

            # vyber s krátkým timeoutem a vyvolej change
            try:
                sel.select_option(value=pick_val, timeout=500)
            except Exception:
                # poslední pokus: změň value přes JS + change
                sel.evaluate("(el, v) => { el.value = v; el.dispatchEvent(new Event('change', {bubbles:true})); }", pick_val)

            page.wait_for_timeout(SLEEP_MS)

            after_txt = (cell.inner_text() or "").strip()
            if after_txt and after_txt != "----":
                if any(_norm_name(after_txt) == _norm_name(v) for v in _name_variants(name)):
                    log(f"  ✓ {name} → {cell_sel or selector} (select)  [after='{after_txt}']")
                else:
                    log(f"  ~ {name} → {cell_sel or selector} vybráno (select), ale zobrazeno '{after_txt}'")
            else:
                log(f"  ⚠ {name} → po selectu žádná změna (stále '{after_txt or ''}')")
            return

        except Exception as e:
            log(f"  ✗ {name} → práce se <select> selhala: {e!r}")
            return

    # ---------- FALLBACK: AUTOCOMPLETE (jen pokud select není) ----------
    ac = cell.locator("input.ui-autocomplete-input, input.ac_input").first
    if not ac.count():
        # poslední možnost: libovolný text input v buňce, ale NE sety
        ac = cell.locator("input[type='text']:not(.zapas-set):not([name^='set'])").first
    if not ac.count():
        log(f"  ✗ {name} → žádný hráčský input/select v buňce ({cell_sel or selector})")
        return

    try:
        try: ac.fill("")
        except Exception: pass
        ac.focus()
        ac.type(name, delay=0)
        page.wait_for_timeout(50)

        menu_sel = "ul.ui-autocomplete:visible, .ui-autocomplete.ui-menu:visible"
        page.wait_for_selector(menu_sel, timeout=MENU_MS)
        menu = page.locator(menu_sel).first.locator("li")

        # rychlý výběr shody
        want_norms = {_norm_name(v) for v in _name_variants(name)}
        pick = -1
        n = min(menu.count(), 20)
        for i in range(n):
            base = _strip_menu_text(menu.nth(i).inner_text() or "")
            if _norm_name(base) in want_norms:
                pick = i; break

        if pick < 0 and " " in name:
            ac.fill(""); ac.focus(); ac.type(name.split()[-1], delay=0)
            page.wait_for_selector(menu_sel, timeout=MENU_MS)
            menu = page.locator(menu_sel).first.locator("li")
            n = min(menu.count(), 20)
            for i in range(n):
                base = _strip_menu_text(menu.nth(i).inner_text() or "")
                if _norm_name(base) in want_norms:
                    pick = i; break

        if pick >= 0:
            menu.nth(pick).click(timeout=CLICK_MS)
            page.wait_for_timeout(SLEEP_MS)
        else:
            log("  žádná shoda v autocomplete – přeskočeno")
            return

        after_txt = (cell.inner_text() or "").strip()
        if after_txt and after_txt != "----":
            if any(_norm_name(after_txt) == _norm_name(v) for v in _name_variants(name)):
                log(f"  ✓ {name} → {cell_sel or selector} (autocomplete)  [after='{after_txt}']")
            else:
                log(f"  ~ {name} → {cell_sel or selector} vybráno (autocomplete), ale zobrazeno '{after_txt}'")
        else:
            log(f"  ⚠ {name} → žádná změna v buňce (stále '{after_txt or ''}')")

    except Exception as e:
        log(f"  ✗ {name} → autocomplete selhal: {e!r}")




def _fill_sets_by_event_index(page, event_index, sets, log):
    """
    Vyplní sety pro daný event podle jeho pozice v seznamu.
    """
    if not sets:
        return
        
    try:
        # Najdi event podle indexu
        event = page.locator(".event").nth(event_index)
        if not event.count():
            log(f"  Event #{event_index} nenalezen")
            return
            
        # Vyplň až 5 setů
        for i, value in enumerate(sets[:5]):
            if not value:
                continue
                
            set_input = event.locator(f".zapas-set[data-set='{i+1}']")
            if set_input.count():
                set_input.fill(str(_map_wo(value)))
                log(f"  set{i+1} ← {value} (event #{event_index})")
            else:
                log(f"  Set {i+1} input nenalezen pro event #{event_index}")
                
    except Exception as e:
        log(f"  Sety pro event #{event_index} selhaly: {repr(e)}")

def _map_wo(val):
    """Mapuje WO značky na STIS kódy."""
    if not val:
        return ""
    s = str(val).strip().upper().replace(" ", "")
    if s in ("WO3:0", "3:0WO"):
        return "101"
    if s in ("WO0:3", "0:3WO"):  
        return "-101"
    return val

def _dom_dump(page, xlsx_path, log):
    try:
        html_path = Path(xlsx_path).with_suffix(".online_dump.html")
        png_path  = Path(xlsx_path).with_suffix(".online_dump.png")
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(page.content())
        page.screenshot(path=str(png_path), full_page=True)
        log(f"DOM dump → {html_path.name}, screenshot → {png_path.name}")
    except Exception as e:
        log(f"DOM dump failed: {repr(e)}")

def cnt(page, css):  # krátká pomůcka do logu
    try:
        return page.locator(css).count()
    except Exception:
        return -1

def fill_online_from_zdroj(page, data, log, xlsx_path=None):
    """
    Vyplní online formulář STIS podle skutečné struktury DOM.

    Očekávaný tvar `data` (z read_zdroj_data):
    {
      "doubles": [
        {"home1": "...", "home2": "...", "away1": "...", "away2": "...", "sets": ["11:7","..."]},  # c0
        {"home1": "...", "home2": "...", "away1": "...", "away2": "...", "sets": [...]},           # c1
      ],
      "singles": [
        {"idx": 2,  "home": "...", "away": "...", "sets": [...]},  # → d0
        ...
        {"idx": 17, "home": "...", "away": "...", "sets": [...]},  # → d15
      ]
    }
    """
    doubles = (data or {}).get("doubles", []) or []
    singles = (data or {}).get("singles", []) or []

    log(
        "fill_online_from_zdroj: start – singles:",
        len(singles),
        "doubles:",
        len(doubles)
    )

    # --- čekej na připravenost online formuláře ---
    try:
        wait_online_ready(page, log)
    except Exception:
        log("Inputs se neobjevily – dělám dump DOMu.")
        if xlsx_path:
            try:
                _dom_dump(page, xlsx_path, log)
            except Exception as e_dump:
                log(f"DOM dump selhal: {e_dump!r}")
        raise

    # --- CSS hack: odhrň překryv kartičky + ujisti viditelnost labelu ---
    page.add_style_tag(content="""
      .button-karta, .button-karta * { pointer-events: none !important; }
      .player-name { visibility: visible !important; opacity: 1 !important; }
    """)
    log("CSS hack pro .button-karta a .player-name aplikován.")

    # ==========================
    # ČTYŘHRA #1 (ID: c0)
    # ==========================
    if len(doubles) >= 1:
        dbl = doubles[0]
        log("Vyplňuji čtyřhru #1 (c0)")

        if dbl.get("home1"):
            log(f"[c0] home1 sel=#c0 .cell-player:first-child  name={dbl['home1']!r}")
            _fill_player_by_click(page, "#c0 .cell-player:first-child", dbl["home1"], log)

        if dbl.get("away1"):
            log(f"[c0] away1 sel=#c0 .cell-player:last-child   name={dbl['away1']!r}")
            _fill_player_by_click(page, "#c0 .cell-player:last-child",  dbl["away1"], log)

        # druhý pár (sourozenec #c0)
        if dbl.get("home2"):
            log(f"[c0] home2 sel=#c0 + .cell-players .cell-player:first-child  name={dbl['home2']!r}")
            _fill_player_by_click(page, "#c0 + .cell-players .cell-player:first-child", dbl["home2"], log)

        if dbl.get("away2"):
            log(f"[c0] away2 sel=#c0 + .cell-players .cell-player:last-child   name={dbl['away2']!r}")
            _fill_player_by_click(page, "#c0 + .cell-players .cell-player:last-child",  dbl["away2"], log)

        if dbl.get("sets"):
            _fill_sets_by_event_index(page, 0, dbl["sets"], log)

    # ==========================
    # ČTYŘHRA #2 (ID: c1)
    # ==========================
    if len(doubles) >= 2:
        dbl = doubles[1]
        log("Vyplňuji čtyřhru #2 (c1)")

        if dbl.get("home1"):
            log(f"[c1] home1 sel=#c1 .cell-player:first-child  name={dbl['home1']!r}")
            _fill_player_by_click(page, "#c1 .cell-player:first-child", dbl["home1"], log)

        if dbl.get("away1"):
            log(f"[c1] away1 sel=#c1 .cell-player:last-child   name={dbl['away1']!r}")
            _fill_player_by_click(page, "#c1 .cell-player:last-child",  dbl["away1"], log)

        # druhý pár (sourozenec #c1)
        if dbl.get("home2"):
            log(f"[c1] home2 sel=#c1 + .cell-players .cell-player:first-child  name={dbl['home2']!r}")
            _fill_player_by_click(page, "#c1 + .cell-players .cell-player:first-child", dbl["home2"], log)

        if dbl.get("away2"):
            log(f"[c1] away2 sel=#c1 + .cell-players .cell-player:last-child   name={dbl['away2']!r}")
            _fill_player_by_click(page, "#c1 + .cell-players .cell-player:last-child",  dbl["away2"], log)

        if dbl.get("sets"):
            _fill_sets_by_event_index(page, 1, dbl["sets"], log)

    # ==========================
    # SINGLY d0..d15 (Excel idx 2..17)
    # ==========================
    for match_data in singles:
        excel_idx = int(match_data.get("idx", 0))  # 2..17
        if excel_idx < 2 or excel_idx > 17:
            continue

        dom_idx   = excel_idx - 2              # 0..15 → #d{dom_idx}
        event_idx = excel_idx                  # pro sety zůstává excelový index

        log(f"Zpracovávám singl Excel#{excel_idx} → DOM d{dom_idx} → event #{event_idx}")

        if match_data.get("home"):
            log(f"[d{dom_idx}] home sel=#d{dom_idx} .cell-player:first-child  name={match_data['home']!r}")
            _fill_player_by_click(page, f"#d{dom_idx} .cell-player:first-child", match_data["home"], log)

        if match_data.get("away"):
            log(f"[d{dom_idx}] away sel=#d{dom_idx} .cell-player:last-child   name={match_data['away']!r}")
            _fill_player_by_click(page, f"#d{dom_idx} .cell-player:last-child",  match_data["away"], log)

        if match_data.get("sets"):
            _fill_sets_by_event_index(page, event_idx, match_data["sets"], log)

    # ==========================
    # Uložit změny
    # ==========================
    log("Klikám 'Uložit změny'…")
    try:
        page.locator("input[name='ulozit']").click(timeout=5000)
        page.wait_for_timeout(1000)
        log("Změny uloženy.")
    except Exception as e:
        log(f"Uložení selhalo: {e!r}")


def read_zdroj_data(xlsx_path, log):
    """
    Vrátí:
    {
      "doubles": [
        {"home1":..,"home2":..,"away1":..,"away2":..,"sets":[...]} ,   # c0
        {"home1":..,"home2":..,"away1":..,"away2":..,"sets":[...]}     # c1
      ],
      "singles": [
        {"idx": 2, "home":.., "away":.., "sets":[...]},   # d0 (Excel ř.7)
        ...
        {"idx":17, "home":.., "away":.., "sets":[...]}    # d15 (Excel ř.22)
      ]
    }
    """
    wb = load_workbook(xlsx_path, data_only=True)
    if "zdroj" not in wb.sheetnames:
        raise RuntimeError(f"V sešitu chybí list 'zdroj'. Máš: {', '.join(wb.sheetnames)}")
    sh = wb["zdroj"]

    log("== DEBUG EXCEL START ==")

    # ---- Doubles #1 (c0) ----
    c0 = {
        "home1": cell_value(sh, "D2"),
        "home2": cell_value(sh, "D3"),
        "away1": cell_value(sh, "E2"),
        "away2": cell_value(sh, "E3"),
        "sets":  row_sets(sh, 3)
    }
    log(f"[EXCEL] c0.home1 D2 = {c0['home1']!r}")
    log(f"[EXCEL] c0.home2 D3 = {c0['home2']!r}")
    log(f"[EXCEL] c0.away1 E2 = {c0['away1']!r}")
    log(f"[EXCEL] c0.away2 E3 = {c0['away2']!r}")
    log(f"[EXCEL] c0.sets  I3–M3 = {c0['sets']}")

    # ---- Doubles #2 (c1) ----
    c1 = {
        "home1": cell_value(sh, "D4"),
        "home2": cell_value(sh, "D5"),
        "away1": cell_value(sh, "E4"),
        "away2": cell_value(sh, "E5"),
        "sets":  row_sets(sh, 5)
    }
    log(f"[EXCEL] c1.home1 D4 = {c1['home1']!r}")
    log(f"[EXCEL] c1.home2 D5 = {c1['home2']!r}")
    log(f"[EXCEL] c1.away1 E4 = {c1['away1']!r}")
    log(f"[EXCEL] c1.away2 E5 = {c1['away2']!r}")
    log(f"[EXCEL] c1.sets  I5–M5 = {c1['sets']}")

    doubles = [c0, c1]

    # ---- Singles d0..d15 (Excel řádky 7..22) ----
    singles = []
    excel_row = 7
    for excel_idx in range(2, 18):  # 2..17
        home = cell_value(sh, f"D{excel_row}")
        away = cell_value(sh, f"E{excel_row}")
        sets = row_sets(sh, excel_row)

        log(f"[EXCEL] d{excel_idx-2}: idx={excel_idx}  D{excel_row}='{home}'  E{excel_row}='{away}'  I{excel_row}-M{excel_row}={sets}")

        singles.append({
            "idx":  excel_idx,   # POZOR: tohle používáme pro event_idx (sety)
            "home": home,
            "away": away,
            "sets": sets
        })
        excel_row += 1

    out = {"doubles": doubles, "singles": singles}

    # Souhrn
    log(f"[EXCEL] SUMMARY doubles: {[(d['home1'],d['home2'],d['away1'],d['away2'],d['sets']) for d in doubles]}")
    log(f"[EXCEL] SUMMARY singles count={len(singles)} first={singles[0]} last={singles[-1]}")
    log("== DEBUG EXCEL END ==")
    return out
        
def boot(msg: str):
    """Zapíš krátkou zprávu ještě před main() – přežije i selhání argparse."""
    line = f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {msg}\n"
    for p in BOOT_FILES:
        try:
            p.parent.mkdir(parents=True, exist_ok=True)
            with open(p, "a", encoding="utf-8") as f:
                f.write(line)
        except Exception:
            pass

def msgbox(text: str, title: str="stis-uploader"):
    try:
        ctypes.windll.user32.MessageBoxW(0, str(text), str(title), 0x40)  # MB_ICONINFORMATION
    except Exception:
        pass

BOOTLOG = Path(os.environ.get("TEMP", str(Path.cwd()))) / "stis_boot.log"

def make_logger(xlsx_path: Path):
    """Vrátí (log_fn, file_handle, log_path) – loguje s časovou značkou."""
    log_path = xlsx_path.with_suffix(".stislog.txt")
    f = open(log_path, "a", encoding="utf-8")
    def log(*parts):
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        line = " ".join(str(p) for p in parts)
        f.write(f"[{ts}] {line}\n")
        f.flush()
    return log, f, log_path

def ensure_pw_browsers(log=None):
    """
    Preferuj prohlížeče v PLAYWRIGHT_BROWSERS_PATH (tj. přibalené u EXE),
    jinak použij uživatelský store %LOCALAPPDATA%\\ms-playwright.
    Když tam Chromium chybí, stáhni ho pomocí `playwright install chromium`
    přímo do téhle cesty.
    """
    default_store = Path(os.environ.get("LOCALAPPDATA", Path.home())) / "ms-playwright"

    # respektuj už nastavené PLAYWRIGHT_BROWSERS_PATH (nastavuje prepare_playwright_browsers)
    store = Path(os.environ.get("PLAYWRIGHT_BROWSERS_PATH") or default_store)
    os.environ["PLAYWRIGHT_BROWSERS_PATH"] = str(store)

    # univerzální detekce: existuje chrome(.exe) kdekoli pod store?
    has_chrome = any(store.rglob("chrome.exe")) or any(store.rglob("chrome"))
    if has_chrome:
        if log: log("Chromium already present in", store)
        return

    if log: log("Chromium not found in", store, "– running: playwright install chromium")

    # jednorázově povol stáhnutí pro explicitní "install"
    import playwright.__main__ as pw_cli
    prev_pw_disable = os.environ.pop("PW_DISABLE_DOWNLOADS", None)
    old_argv = sys.argv[:]
    try:
        store.mkdir(parents=True, exist_ok=True)
        os.environ["PLAYWRIGHT_BROWSERS_PATH"] = str(store)
        sys.argv = ["playwright", "install", "chromium"]
        pw_cli.main()
        if log: log("Playwright install finished.")
    finally:
        sys.argv = old_argv
        # po instalaci zase zablokuj automatické stahování
        os.environ["PLAYWRIGHT_BROWSERS_PATH"] = str(store)
        os.environ["PW_DISABLE_DOWNLOADS"] = prev_pw_disable or "1"




def norm(x) -> str:
    s = "" if x is None else str(x)
    s = s.strip().lower()
    # odstraň diakritiku
    s = "".join(ch for ch in unicodedata.normalize("NFD", s)
                if unicodedata.category(ch) != "Mn")
    # bez mezer a nealfanumerických znaků
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[^0-9a-z_]", "", s)
    return s

def as_time_txt(v):
    import re, datetime
    if v is None:
        return None

    # 1) Excel přes openpyxl často vrátí datetime.time
    if isinstance(v, datetime.time):
        return f"{v.hour:02d}:{v.minute:02d}"

    # 2) Excel čas jako frakce dne (0.0–1.0)
    if isinstance(v, (int, float)):
        if 0 <= v < 2:  # tolerantní (někdy bývá 1.0 + zlomky)
            total_minutes = int(round((v % 1) * 24 * 60))
            hh = (total_minutes // 60) % 24
            mm = total_minutes % 60
            return f"{hh:02d}:{mm:02d}"

    # 3) Textové varianty: "19", "19:0", "19:00", "19.00", "19 00", "7 pm"
    if isinstance(v, str):
        s = v.strip().lower()

        # "7 pm", "7:30 am"
        m = re.match(r'^(\d{1,2})(?::(\d{1,2}))?\s*(am|pm)$', s)
        if m:
            h = int(m.group(1)); m2 = int(m.group(2) or 0)
            ap = m.group(3)
            if ap == 'pm' and h != 12: h += 12
            if ap == 'am' and h == 12: h = 0
            if 0 <= h <= 23 and 0 <= m2 <= 59:
                return f"{h:02d}:{m2:02d}"

        # "19", "19:0", "19:00", "19.00", "19 00", "19,00"
        m = re.match(r'^(\d{1,2})[ :\.\,h]?(\d{0,2})$', s)
        if m:
            h = int(m.group(1))
            m2 = int(m.group(2)) if m.group(2) else 0
            if 0 <= h <= 23 and 0 <= m2 <= 59:
                return f"{h:02d}:{m2:02d}"

    return None


def get_setup_sheet(wb):
    """Najdi list 'setup' case-insensitive, jinak vrať první list."""
    for ws in wb.worksheets:
        if (ws.title or "").strip().lower() == "setup":
            return ws
    return wb.active

def find_login_pwd(ws):
    """Vrátí login/heslo. Nejprve z B1/B2, když chybí, zkusí popisky 'login' / 'heslo' v horní části listu."""
    login = str(ws["B1"].value or "").strip()
    pwd   = str(ws["B2"].value or "").strip()
    if login and pwd:
        return login, pwd

    max_r = min(30, ws.max_row or 0)
    max_c = min(20, ws.max_column or 0)
    found_login = found_pwd = ""
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = norm(ws.cell(r, c).value or "")
            if v == "login" and c + 1 <= (ws.max_column or 0):
                found_login = str(ws.cell(r, c + 1).value or "").strip()
            if v == "heslo" and c + 1 <= (ws.max_column or 0):
                found_pwd = str(ws.cell(r, c + 1).value or "").strip()
    return found_login, found_pwd

def find_teams_header_anywhere(wb):
    """
    Projdi všechny listy a najdi první řádek, kde je hlavička tabulky Teams:
      - obsahuje 'druzstvo' (ale ne 'id' / 'vedouci')
      - a zároveň některý JASNÝ marker ID ('druzstvoid', 'id_druzstva', 'iddruzstva') nebo PŘESNĚ 'id'
    Vrací (sheet, hdr_row) nebo (None, None).
    """
    strict_id_markers = ("druzstvoid", "id_druzstva", "iddruzstva")  # žádné volné "id" jako podřetězec!
    for ws in wb.worksheets:
        max_r = min(60, ws.max_row or 0)
        max_c = min(80, ws.max_column or 0)
        for r in range(1, max_r + 1):
            row_norm = [norm(ws.cell(r, c).value or "") for c in range(1, max_c + 1)]
            has_name = any(("druzstvo" in v and "id" not in v and "vedouci" not in v) for v in row_norm)
            has_id   = any(any(m in v for m in strict_id_markers) or (v == "id") for v in row_norm)
            if has_name and has_id:
                return ws, r
    return None, None

def open_match_form(page, log):
    """Na stránce družstva otevře formulář – preferuje 'vložit zápis', jinak 'upravit zápis'.
       Zkouší text i href (zapis_start.php / online.php). Vrací True/False.
    """
    try:
        page.wait_for_selector(
            "a:has-text('vložit zápis'), a:has-text('upravit zápis'), "
            "a[href*='zapis_start.php?u='], a[href*='online.php?u=']",
            timeout=15000
        )
    except Exception:
        log("Nenalezl jsem žádný z očekávaných odkazů do 15 s.")
        return False

    candidates = [
        page.get_by_role("link", name=re.compile(r"vložit\s*zápis", re.I)),
        page.get_by_role("link", name=re.compile(r"upravit\s*zápis", re.I)),
        page.locator("a[href*='zapis_start.php?u=']"),
        page.locator("a[href*='online.php?u=']"),
    ]

    for i, sel in enumerate(candidates, 1):
        try:
            if sel.count():
                log(f"Zkouším selector {i}")
                sel.first.click(timeout=5000)
                page.wait_for_load_state("domcontentloaded")

                if page.locator("text=/vkládání zápisu/i").count() \
                   or "online.php?u=" in page.url \
                   or "zapis_start.php?u=" in page.url:
                    log("Formulář otevřen na URL:", page.url)
                    return True

                if page.locator("text=/špatn.*url/i").count():
                    log("Server hlásí 'špatné URL' – zkusím jiný odkaz.")
                    page.go_back()
                    page.wait_for_load_state("domcontentloaded")
        except Exception as e:
            log(f"Selector {i} selhal:", repr(e))

    # poslední záchrana – proskenuj všechny <a> a skoč přímo na vyhovující href
    try:
        anchors = page.eval_on_selector_all(
            "a",
            "els => els.map(a => ({text: (a.textContent||'').trim(), href: a.href||''}))"
        )
        for a in anchors:
            if re.search(r"(zapis_start\.php|online\.php)\?u=\d+", a.get("href",""), re.I):
                log("Jdu přímo na", a["href"])
                page.goto(a["href"], wait_until="domcontentloaded")
                return True
    except Exception as e:
        log("Fallback scan anchorů selhal:", repr(e))

    return False

def read_excel_config(xlsx_path: Path, team_name: str):
    """
    Najde list s tabulkou Teams kdekoli v sešitu, přihlášení bere z B1/B2
    (nebo z popisků 'login'/'heslo'), namapuje sloupce a vrátí login, heslo a dict týmu.
    """
    wb = load_workbook(xlsx_path, data_only=True)

    # 1) Najdi list a řádek hlavičky Teams kdekoli v sešitu
    setup, hdr_row = find_teams_header_anywhere(wb)
    if setup is None:
        # diagnostika – co jsme nahoře viděli
        try:
            with open(Path(xlsx_path).with_suffix(".log"), "w", encoding="utf-8") as f:
                f.write("Header not found. Top rows (normalized):\n")
                for ws in wb.worksheets:
                    f.write(f"[{ws.title}]\n")
                    max_r = min(15, ws.max_row or 0)
                    max_c = min(20, ws.max_column or 0)
                    for r in range(1, max_r + 1):
                        rn = [norm(ws.cell(r, c).value or "") for c in range(1, max_c + 1)]
                        f.write(f"R{r}: {rn}\n")
        except Exception:
            pass
        raise RuntimeError("V setup/Teams chybí sloupce 'Družstvo' a/nebo 'DruzstvoID'.")

    # 2) Login a heslo (z nalezeného listu)
    login, pwd = find_login_pwd(setup)
    if not login or not pwd:
        raise RuntimeError("Vyplň login/heslo (B1/B2, nebo vedle popisků 'login'/'heslo').")

    # 3) Namapuj sloupce podle hlavičky (ID sloupec detekuj STRIKTNĚ + vyluč 'vedouci*')
    max_c = min(80, setup.max_column or 0)
    idx = {}
    for c in range(1, max_c + 1):
        h = norm(setup.cell(hdr_row, c).value or "")

        # Družstvo (název)
        if ("druzstvo" in h) and ("id" not in h) and ("vedouci" not in h):
            idx["name"] = c

        # ID družstva – akceptuj jasné varianty a/nebo přesné 'id', nikdy ne cokoliv s 'vedouci'
        is_id_col = (("druzstvoid" in h) or ("id_druzstva" in h) or ("iddruzstva" in h) or (h == "id")) \
                    and ("vedouci" not in h)
        if is_id_col:
            idx["id"] = c

        # Vedoucí domácích / hostů
        if "vedoucidomacich" in h or ("vedouci" in h and "host" not in h):
            idx["ved_dom"] = c
        if "vedoucihostu" in h or ("vedouci" in h and "host" in h):
            idx["ved_host"] = c

        # Herna, začátek, konec
        if "herna" in h:
            idx["herna"] = c
        if "zacatekut" in h or "zacatek" in h:
            idx["zacatek"] = c
        if "konecutkani" in h or "konec" in h:
            idx["konec"] = c

    if "name" not in idx or "id" not in idx:
        raise RuntimeError("V Teams chybí sloupce 'Družstvo' a/nebo 'DruzstvoID'.")

    # 4) Najdi požadované družstvo pod hlavičkou
    team = None
    r = hdr_row + 1
    while r <= (setup.max_row or 0):
        nm = setup.cell(r, idx["name"]).value
        if nm is None or str(nm).strip() == "":
            break
        if str(nm).strip().lower() == team_name.strip().lower():
            def getcol(key):
                c = idx.get(key)
                return setup.cell(r, c).value if c else None
            raw_z = getcol("zacatek")
            raw_k = getcol("konec")
            team = {
                "name":    str(nm).strip(),
                "id":      str(getcol("id") or "").strip(),
                "ved_dom": getcol("ved_dom"),
                "ved_host":getcol("ved_host"),
                "herna":   getcol("herna"),
                # NOVÉ: uchovej RAW hodnoty + parsed řetězec HH:MM
                "zacatek_raw": raw_z,
                "konec_raw":   raw_k,
                "zacatek":     as_time_txt(raw_z),
                "konec":       as_time_txt(raw_k),
            }
            break
        r += 1

    if not team:
        raise RuntimeError(f"Družstvo '{team_name}' nenalezeno v Teams.")
    if not team["id"]:
        raise RuntimeError("Prázdné DruzstvoID u zvoleného družstva.")

    # 5) Očisti ID na čisté číslo (zabráníme tomu, aby se do URL dostalo 'Kozel Petr' apod.)
    raw_id = str(team.get("id", "")).strip()
    m = re.search(r"\d+", raw_id)
    if not m:
        raise RuntimeError(f"Neplatné DruzstvoID: {raw_id!r}")
    team["id"] = m.group(0)

    return login, pwd, team

def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", required=True, help="plná cesta k XLSX")
    p.add_argument("--team", required=True, help="název družstva (sloupec 'Družstvo')")
    g = p.add_mutually_exclusive_group()
    g.add_argument("--headed",  dest="headed",  action="store_true",  help="viditelný prohlížeč")
    g.add_argument("--headless", dest="headed", action="store_false", help="bez UI")
    p.set_defaults(headed=True)  # výchozí = viditelné okno
    return p.parse_args()

# Nahraďte celou main() funkci tímto opraveným kódem:

def main():
    args = parse_args()
    xlsx_path = Path(args.xlsx).resolve()
    if not xlsx_path.exists():
        raise RuntimeError(f"Soubor neexistuje: {xlsx_path}")

    # logger vedle XLSX
    log, log_file, log_path = make_logger(xlsx_path)
    log("==== stis_uploader start ====")
    log("XLSX:", xlsx_path)
    log("Team:", args.team)
    log("Headed:", getattr(args, "headed", True))

    try:
        # 1) načti přihlášení + tým
        user_login, user_pwd, team = read_excel_config(xlsx_path, args.team)
        log("Login OK; team:", team["name"], "ID:", team["id"])
        log("Time (XLSX raw → parsed):", repr(team.get("zacatek_raw")), "→", team.get("zacatek"))

        # 1.5) data ze "zdroj"
        try:
            zdroj_data = read_zdroj_data(xlsx_path, log)  # NOVÁ signatura s log
        except TypeError:
            log("WARNING: read_zdroj_data(xlsx_path, log) není k dispozici – volám starou verzi bez logování.")
            zdroj_data = read_zdroj_data(xlsx_path)
        except Exception as e:
            log("WARNING: Nepodařilo se načíst data ze 'zdroj' listu:", repr(e))
            zdroj_data = None

        if not zdroj_data:
            log("WARNING: zdroj_data=None → nebude se vybírat žádný hráč (vyplní se jen sety, pokud jsou).")
        else:
            dbls = zdroj_data.get("doubles", []) or []
            sgls = zdroj_data.get("singles", []) or []
            log(f"Zdroj data loaded – doubles: {len(dbls)}, singles: {len(sgls)}")
            for i, d in enumerate(dbls):
                log(f"[EXCEL] c{i}: home1={d.get('home1')!r}, home2={d.get('home2')!r}, "
                    f"away1={d.get('away1')!r}, away2={d.get('away2')!r}, sets={d.get('sets')}")
            for s in sgls:
                idx = s.get("idx")
                dom_idx = (idx or 2) - 2
                log(f"[EXCEL] d{dom_idx}: idx={idx} home={s.get('home')!r} away={s.get('away')!r} sets={s.get('sets')}")

        headed   = bool(getattr(args, "headed", True))
        headless = not headed

        # 2) připrav Playwright runtime
        prepare_playwright_browsers(log)   # nastaví PLAYWRIGHT_BROWSERS_PATH
        ensure_pw_browsers(log)            # případně doinstaluje Chromium

        with sync_playwright() as p:
            # 3) spuštění prohlížeče (Chromium → Chrome → Edge)
            log("Launching browser… headless =", headless)
            browser = None
            try:
                browser = p.chromium.launch(headless=headless)
                log("Launched: managed Chromium")
            except Exception as e1:
                log("Chromium failed:", repr(e1), "→ trying channel=chrome")
                try:
                    browser = p.chromium.launch(channel="chrome", headless=headless)
                    log("Launched: channel=chrome")
                except Exception as e2:
                    log("Chrome failed:", repr(e2), "→ trying channel=msedge")
                    browser = p.chromium.launch(channel="msedge", headless=headless)
                    log("Launched: channel=msedge")

            context = browser.new_context()
            page = context.new_page()

            # DŮLEŽITÉ: krátký default timeout (žádné 30s visení)
            page.set_default_timeout(1500)

            # 4) login
            log("Navigating to login…")
            page.goto("https://registr.ping-pong.cz/htm/auth/login.php", wait_until="domcontentloaded", timeout=20000)
            page.fill("input[name='login']", user_login)
            page.fill("input[name='heslo']",  user_pwd)
            
            btn_login = page.locator("[name='send']")
            # Login vyvolává navigaci → dej mu delší timeout jen tady
            with page.expect_navigation(wait_until="domcontentloaded", timeout=15000):
                btn_login.click(timeout=3000)
            log("Logged in.")

            # 5) stránka družstva
            team_url = f"https://registr.ping-pong.cz/htm/auth/klub/druzstva/vysledky/?druzstvo={team['id']}"
            log("Open team page:", team_url)
            page.goto(team_url, wait_until="domcontentloaded", timeout=20000)


            # 6) najdi vstup do formuláře (vložit/upravit)
            log("Hledám odkaz 'vložit/upravit zápis'…")
            if not open_match_form(page, log):
                raise RuntimeError("Na stránce družstva jsem nenašel odkaz do formuláře.")

            # 7.1) Hrací místnost (SELECT podle labelu „Hrací místnost“)
            wanted_room_text = (team.get("hraci_mistnost") or team.get("herna") or "").strip()
            ok_room = fill_playroom(page, wanted_text=wanted_room_text, log=log)

            log(f"Hrací místnost → {'OK' if ok_room else 'NEVYBRÁNA'}")

            # 7.2) Začátek utkání (hh:mm)
            start_txt = (team.get("zacatek") or "19:00").strip()
            try:
                if ":" in start_txt:
                    hh, mm = start_txt.split(":")[:2]
                else:
                    hh, mm = "19", "00"
                hh = int(hh); mm = int(mm)

                if page.locator("select[name='zapis_zacatek_hodiny']").count():
                    page.select_option("select[name='zapis_zacatek_hodiny']", value=str(hh))
                    log(f"Hodina nastavena: {hh:02d}")
                if page.locator("select[name='zapis_zacatek_minuty']").count():
                    page.select_option("select[name='zapis_zacatek_minuty']", value=str(mm))
                    log(f"Minuta nastavena: {mm:02d}")

                # vystřel change na obou selectech
                try:
                    page.evaluate("document.querySelector('select[name=\"zapis_zacatek_hodiny\"]').dispatchEvent(new Event('change',{bubbles:true}))")
                    page.evaluate("document.querySelector('select[name=\"zapis_zacatek_minuty\"]').dispatchEvent(new Event('change',{bubbles:true}))")
                except Exception:
                    pass
                page.wait_for_timeout(200)
                log("Začátek nastaven:", f"{hh:02d}:{mm:02d}")
            except Exception as e:
                log("Set start time failed:", repr(e))
                try:
                    page.select_option("select[name='zapis_zacatek_hodiny']", value="19")
                    page.select_option("select[name='zapis_zacatek_minuty']", value="0")
                    log("Fallback čas: 19:00")
                except Exception:
                    pass

            # 7.3) Vedoucí družstev (autocomplete → vybrat položku z menu)
            leaders_only_from_club = bool(team.get("leaders_only_from_club", True))
            ok_leaders = fill_leaders_on_start(
                page,
                home_name_text=str(team.get("ved_dom_text") or team.get("ved_dom") or "").strip(),
                away_name_text=str(team.get("ved_host_text") or team.get("ved_host") or "").strip(),
                log=log,
                only_from_club=True,
            )

            log(f"Vedoucí → {'OK' if ok_leaders else 'NEULOŽENO'}")

            # 8) Odeslat úvodní formulář (s malým retry na chybovou hlášku času)
            max_attempts = 3
            for attempt in range(max_attempts):
                log(f"Pokus {attempt+1}/{max_attempts}: Click 'Uložit a pokračovat'…")
                try:
                    btn = page.locator("input[name='odeslat']")
                    if btn.count():
                        # tento klik také naviguje → expect_navigation s delším timeoutem
                        with page.expect_navigation(wait_until="domcontentloaded", timeout=15000):
                            btn.click(timeout=3000)
                        log("Formulář odeslán")
            
                        # kontrola chybové hlášky k času (už na nové stránce)
                        if page.locator(".exception:has-text('není vyplněn začátek utkání')").count():
                            log(f"Pokus {attempt+1}: Server stále hlásí chybu s časem")
                            if attempt < max_attempts - 1:
                                page.select_option("select[name='zapis_zacatek_hodiny']", value=str(hh))
                                page.select_option("select[name='zapis_zacatek_minuty']", value=str(mm))
                                page.wait_for_timeout(200)
                                continue
                        break
                except Exception as e:
                    log(f"Pokus {attempt+1} selhal:", repr(e))
                    if attempt == max_attempts - 1:
                        raise RuntimeError("Nepodařilo se odeslat formulář ani po několika pokusech")


            # 9) Čekej na online editor a vyplň
            try:
                page.wait_for_function(
                    "window.location.href.includes('online.php') || document.querySelector('input.zapas-set') !== null",
                    timeout=30000
                )
                log("Online editor dostupný na:", page.url)

                # krátký default timeout i pro online část
                page.set_default_timeout(1500)

                if zdroj_data:
                    log("Začínám vyplňovat sestavy a sety…")
                    fill_online_from_zdroj(page, zdroj_data, log, xlsx_path)
                    log("Sestavy a sety vyplněny")
                else:
                    log("VAROVÁNÍ: Žádná data ze 'zdroj' listu k vyplnění")
            except Exception as e:
                log("Problém s online editorem:", repr(e))
                if xlsx_path:
                    _dom_dump(page, xlsx_path, log)
                raise

            # 10) Ukončení
            if headed:
                log("=" * 60)
                log("HOTOVO! Okno prohlížeče zůstává otevřené.")
                log("Zkontrolujte vyplněná data a ručně zavřete okno prohlížeče.")
                log("Program se ukončí až po zavření okna.")
                log("=" * 60)
                try:
                    page.wait_for_event("close", timeout=0)
                    log("Okno prohlížeče bylo zavřeno uživatelem.")
                except Exception as e:
                    log("Čekání na zavření okna skončilo:", repr(e))
                finally:
                    try:
                        context.close(); log("Browser context uzavřen.")
                    except Exception:
                        pass
                    try:
                        browser.close(); log("Browser uzavřen.")
                    except Exception:
                        pass
            else:
                log("Headless režim – zavírám browser automaticky.")
                try: context.close()
                except Exception: pass
                try: browser.close()
                except Exception: pass

    except Exception as e:
        log("ERROR:", repr(e))
        log(traceback.format_exc())
        try:
            os.startfile(str(log_path))
        except Exception:
            pass
        raise
    finally:
        try:
            log("==== stis_uploader end ====")
            log_file.close()
        except Exception:
            pass

if __name__ == "__main__":
    boot("=== EXE start ===")
    try:
        boot("argv: " + " ".join(sys.argv))
        main()
        boot("main() finished OK")
    except SystemExit as e:
        boot(f"SystemExit (pravděpodobně argparse): code={getattr(e, 'code', None)}")
        msgbox("Spuštění skončilo hned na začátku (špatné/neúplné argumenty?).\n" +
               "Zkontroluj prosím volání z Excelu.\n" +
               "V TEMP nebo vedle EXE je stis_boot.log s detaily.")
        raise
    except Exception as e:
        boot("CRASH: " + repr(e))
        try:
            boot(traceback.format_exc())
        except Exception:
            pass
        msgbox(f"Nastala chyba: {e}\nPodrobnosti jsou ve stis_boot.log.")
        raise
    finally:
        boot("=== EXE end ===")
